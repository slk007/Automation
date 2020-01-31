import openpyxl

# path
path = "/media/shubham/WORK/Pythontest.xlsx"

# object for workbook
workbook = openpyxl.load_workbook(path)

# object for sheet
sheet = workbook.active

# title of sheet
print("Title:", sheet.title)
# sheet.title = "Load Extract"
# print("Title:", sheet.title)

# number of rows
maxrows = sheet.max_row
print("No of rows", maxrows)

# number of columns
maxcolumns = sheet.max_column
print("No of columns", maxcolumns)

for i in range(1, maxrows+1):
    cell = sheet.cell(row=i, column=6)  # object for cell in a sheet
    if cell.value is not None:
        print(cell.value)   # getting value for cell

# printing all column names
for i in range(1, maxcolumns+1):
    cell = sheet.cell(row=4, column=i)
    if cell.value is not None:
        print(cell.value)

# # deleting first column
# sheet.delete_cols(1, 1)
#
# # deleting first column
# sheet.delete_rows(1, 3)
#
# print("After deleting the rows and columns")
#
# # number of rows
# maxrows = sheet.max_row
# print("No of rows", maxrows)
#
# # number of columns
# maxcolumns = sheet.max_column
# print("No of columns", maxcolumns)
#
# for i in range(1, maxrows+1):
#     cell = sheet.cell(row=i, column=6)  # object for cell in a sheet
#     if cell.value is not None:
#         print(cell.value)   # getting value for cell
#
# # printing all column names
# for i in range(1, maxcolumns+1):
#     cell = sheet.cell(row=4, column=i)
#     if cell.value is not None:
#         print(cell.value)

# inserting a column
# sheet.insert_cols(6, 1)

for r in range(1, sheet.max_row):
    cell_left = sheet.cell(row=r, column=5)
    s = str(cell_left.value)
    changed = []
    for i in range(0, len(s)):
        if s[i] is '.':
            changed.append(",")
        elif s[i] is ',':
            changed.append(".")
        elif s[i] is '-':
            continue
        else:
            changed.append(s[i])
    cell_right = sheet.cell(row=r, column=6)
    new_value = "".join(map(str, changed))
    cell_right.value = new_value
workbook.save(path)
